from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.models import Registration

TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=12)
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="111827")  # quase preto (dark)
_THIN_SIDE = Side(style="thin", color="2D3748")
BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)

REGISTRATION_COLUMNS = [
    "ID", "Nome", "CPF", "Telefone", "IAP", "Transporte", "Lote",
    "Valor Lote (R$)", "Tipo Pagamento", "Parcelas", "Status", "Mensagem Status",
    "Tem filhos até 5?", "Nomes filhos até 5", "Idade", "Membro da igreja?",
    "Aceitou sem reembolso?", "Comprovante Pix enviado?", "Comprovante enviado em",
    "Revisado por (user_id)", "Revisado em", "Nota revisão", "Criado em", "Atualizado em",
]

REGISTRATION_COLUMN_WIDTHS = {
    "A": 8,   "B": 30, "C": 16, "D": 16, "E": 24, "F": 12,
    "G": 12,  "H": 14, "I": 14, "J": 10, "K": 22, "L": 30,
    "M": 16,  "N": 32, "O": 8,  "P": 18, "Q": 22, "R": 22,
    "S": 18,  "T": 18, "U": 18, "V": 28, "W": 18, "X": 18,
}


# ===== formatação =====

def _fmt_dt(dt) -> str:
    return dt.strftime("%d/%m/%Y %H:%M") if dt else ""


def _yn(value) -> str:
    return "SIM" if value else "NÃO"


def _money_from_cents(cents) -> float:
    return float((cents or 0) / 100.0)


def _safe_text(value):
    return (value or "").strip() if isinstance(value, str) else (value if value is not None else "")


def _parse_date(value: str):
    try:
        return datetime.strptime(value, "%Y-%m-%d")
    except (TypeError, ValueError):
        return None


def _count_by(items, key_fn) -> dict:
    counts: dict = {}
    for item in items:
        key = key_fn(item)
        counts[key] = counts.get(key, 0) + 1
    return counts


def _sort_count_dict(counts: dict) -> list:
    return sorted(counts.items(), key=lambda pair: (-pair[1], str(pair[0])))


# ===== estilo =====

def _style_header_row(worksheet, row: int = 1) -> None:
    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
        cell.border = BORDER


def _apply_table_style(worksheet, start_row, start_col, end_row, end_col, freeze_panes="A2", auto_filter=True) -> None:
    worksheet.freeze_panes = freeze_panes
    if auto_filter:
        worksheet.auto_filter.ref = (
            f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        )
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _set_widths(worksheet, widths: dict) -> None:
    for col_letter, width in widths.items():
        worksheet.column_dimensions[col_letter].width = width


def _make_count_sheet(workbook, title, pairs, col1="Categoria", col2="Qtd"):
    sheet = workbook.create_sheet(title)
    sheet.append([col1, col2])
    _style_header_row(sheet, 1)
    for key, value in pairs:
        sheet.append([key, value])
    _set_widths(sheet, {"A": 36, "B": 10})
    _apply_table_style(sheet, 1, 1, sheet.max_row, 2, freeze_panes="A2", auto_filter=True)
    return sheet


# ===== consulta =====

def _query_registrations(status: str, payment_type: str, date_from: str, date_to: str):
    query = Registration.query

    if status:
        query = query.filter(Registration.status == status)
    if payment_type:
        query = query.filter(Registration.payment_type == payment_type)

    parsed_from = _parse_date(date_from)
    parsed_to = _parse_date(date_to)
    if parsed_from:
        query = query.filter(Registration.created_at >= parsed_from)
    if parsed_to:
        # inclui o dia final (até 23:59:59)
        query = query.filter(
            Registration.created_at < datetime(parsed_to.year, parsed_to.month, parsed_to.day, 23, 59, 59)
        )

    return query.order_by(Registration.created_at.desc()).all()


def _describe_filters(status: str, payment_type: str, date_from: str, date_to: str) -> list[str]:
    filtros = []
    if status:
        filtros.append(f"status={status}")
    if payment_type:
        filtros.append(f"payment_type={payment_type}")
    if date_from:
        filtros.append(f"from={date_from}")
    if date_to:
        filtros.append(f"to={date_to}")
    return filtros


# ===== abas =====

def _build_resumo_sheet(workbook, regs, filtros, by_status, by_pay, by_kids, by_proof) -> None:
    ws = workbook.create_sheet("Resumo")
    ws["A1"] = "Relatório de Inscrições — Tempo de Resplandecer"
    ws["A1"].font = TITLE_FONT

    ws["A3"] = "Gerado em:"
    ws["B3"] = datetime.now().strftime("%d/%m/%Y %H:%M")

    ws["A4"] = "Filtros:"
    ws["B4"] = ", ".join(filtros) if filtros else "—"

    ws["A6"] = "KPIs"
    ws["A6"].font = SECTION_FONT

    kpi_rows = [
        ("Total de inscrições", len(regs)),
        ("Confirmadas", by_status.get("CONFIRMADA", 0)),
        ("Aguardando confirmação", by_status.get("AGUARDANDO_CONFIRMACAO", 0)),
        ("Negadas", by_status.get("NEGADA", 0)),
        ("Pagamentos Pix", by_pay.get("Pix", 0)),
        ("Pagamentos Crédito", by_pay.get("Crédito", 0)),
        ("Tem filhos até 5 (SIM)", by_kids.get("SIM", 0)),
        ("Tem filhos até 5 (NÃO)", by_kids.get("NÃO", 0)),
        ("Pix com comprovante (SIM)", by_proof.get("SIM", 0)),
        ("Pix sem comprovante (NÃO)", by_proof.get("NÃO", 0)),
    ]

    ws["A7"] = "Métrica"
    ws["B7"] = "Valor"
    for coord in ("A7", "B7"):
        ws[coord].font = HEADER_FONT
        ws[coord].fill = HEADER_FILL
        ws[coord].border = BORDER

    row = 8
    for name, value in kpi_rows:
        ws.cell(row=row, column=1, value=name).border = BORDER
        ws.cell(row=row, column=2, value=value).border = BORDER
        row += 1

    _set_widths(ws, {"A": 36, "B": 16})


def _build_inscritos_sheet(workbook, regs) -> None:
    ws = workbook.create_sheet("Inscritos")
    ws.append(REGISTRATION_COLUMNS)
    _style_header_row(ws, 1)

    for r in regs:
        ws.append([
            r.id,
            r.full_name,
            r.cpf,
            r.phone,
            r.iap_local,
            "Ônibus" if r.transport == "onibus" else "Carro",
            r.lot_name,
            _money_from_cents(r.lot_value_cents),
            "Pix" if r.payment_type == "pix" else "Crédito",
            int(r.installments or 1),
            r.status,
            _safe_text(r.status_message),
            _yn(r.has_kids_u5),
            _safe_text(r.kids_u5_names),
            r.age if r.age is not None else "",
            _yn(r.is_church_member),
            _yn(r.agree_no_refund),
            ("SIM" if r.proof_file_path else "NÃO") if r.payment_type == "pix" else "N/A",
            _fmt_dt(r.proof_uploaded_at) if r.payment_type == "pix" else "",
            r.reviewed_by_user_id or "",
            _fmt_dt(r.reviewed_at),
            _safe_text(r.review_note),
            _fmt_dt(r.created_at),
            _fmt_dt(r.updated_at),
        ])

    # formata coluna de dinheiro
    for row in range(2, ws.max_row + 1):
        ws[f"H{row}"].number_format = "0.00"

    _set_widths(ws, REGISTRATION_COLUMN_WIDTHS)
    _apply_table_style(ws, 1, 1, ws.max_row, ws.max_column, freeze_panes="A2", auto_filter=True)


def _build_pagamentos_sheet(workbook, by_pay: dict, by_inst: dict) -> None:
    ws = workbook.create_sheet("Pagamentos")
    ws.append(["Tipo", "Qtd"])
    ws.append(["Pix", by_pay.get("Pix", 0)])
    ws.append(["Crédito", by_pay.get("Crédito", 0)])
    _style_header_row(ws, 1)
    _apply_table_style(ws, 1, 1, ws.max_row, 2, freeze_panes="A2", auto_filter=False)
    _set_widths(ws, {"A": 20, "B": 10})

    # tabela por parcelas
    ws["D1"] = "Parcelas"
    ws["D1"].font = SECTION_FONT

    ws.cell(row=2, column=4, value="Parcelas").font = HEADER_FONT
    ws.cell(row=2, column=5, value="Qtd").font = HEADER_FONT
    ws.cell(row=2, column=4).fill = HEADER_FILL
    ws.cell(row=2, column=5).fill = HEADER_FILL
    ws.cell(row=2, column=4).border = BORDER
    ws.cell(row=2, column=5).border = BORDER

    row = 3
    for key, value in _sort_count_dict(by_inst):
        ws.cell(row=row, column=4, value=key).border = BORDER
        ws.cell(row=row, column=5, value=value).border = BORDER
        row += 1

    _set_widths(ws, {"D": 14, "E": 10})


# ===== entrypoint =====

def build_registrations_workbook(*, status: str = "", payment_type: str = "", date_from: str = "", date_to: str = "") -> BytesIO:
    regs = _query_registrations(status, payment_type, date_from, date_to)

    by_status = _count_by(regs, lambda r: _safe_text(r.status))
    by_pay = _count_by(regs, lambda r: "Pix" if r.payment_type == "pix" else "Crédito")
    by_inst = _count_by(regs, lambda r: f"{int(r.installments or 1)}x")
    by_kids = _count_by(regs, lambda r: "SIM" if r.has_kids_u5 else "NÃO")
    by_transport = _count_by(regs, lambda r: "Ônibus" if r.transport == "onibus" else "Carro")
    by_proof = _count_by(
        [r for r in regs if r.payment_type == "pix"],
        lambda r: "SIM" if r.proof_file_path else "NÃO",
    )
    by_iap = _count_by(regs, lambda r: _safe_text(r.iap_local) or "—")
    by_lot = _count_by(regs, lambda r: _safe_text(r.lot_name) or "—")

    workbook = Workbook()
    workbook.remove(workbook.active)

    _build_resumo_sheet(workbook, regs, _describe_filters(status, payment_type, date_from, date_to), by_status, by_pay, by_kids, by_proof)
    _build_inscritos_sheet(workbook, regs)
    _make_count_sheet(workbook, "Status", _sort_count_dict(by_status), "Status", "Qtd")
    _build_pagamentos_sheet(workbook, by_pay, by_inst)
    _make_count_sheet(workbook, "Kids U5", _sort_count_dict(by_kids), "Tem filhos até 5?", "Qtd")
    _make_count_sheet(workbook, "IAP", _sort_count_dict(by_iap), "IAP", "Qtd")
    _make_count_sheet(workbook, "Lotes", _sort_count_dict(by_lot), "Lote", "Qtd")
    _make_count_sheet(workbook, "Transporte", _sort_count_dict(by_transport), "Transporte", "Qtd")

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
