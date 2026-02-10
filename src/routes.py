from decimal import ROUND_DOWN, ROUND_HALF_UP, Decimal
import os
from datetime import datetime
import uuid
from flask import jsonify, render_template, request, redirect, send_file, url_for, flash, abort, Response
from flask_login import login_user, logout_user, login_required, current_user
from werkzeug.utils import secure_filename

import qrcode
from io import BytesIO
from src.controllers.pix_emv import build_pix_payload
from src import app, database, get_current_lot_info, split_installments
from src.models import AppSetting, User, Role, Registration, AuditLog
from src.forms import ChangePasswordForm, RegisterAndSignupForm, LoginForm, UploadProofForm, ReviewRegistrationForm
from src.decorators import admin_required, payment_reviewer_required, super_required
from src.controllers.b2_utils import upload_to_b2
import secrets
import string
from sqlalchemy import func, or_

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

PIX_PADRAO_MSG = (
    "Informamos que todos os pagamentos realizados via Pix — seja em valor integral ou parcelado — "
    "devem conter obrigatoriamente os centavos finalizados em 0,09. "
    "Essa padronização é necessária para a correta identificação do pagamento. Agradecemos a compreensão."
)

CRIANCAS_MSG = "Crianças de até 5 anos não pagam a inscrição, desde que dividam a cama com o responsável."

INCLUI_ITENS = [
    "Dia 20 — Almoço e jantar",
    "Dia 21 — Café da manhã, almoço e jantar",
    "Dia 22 — Café da manhã",
    "Transporte de ônibus (caso prefira) — Saída de Manaus",
    "Quarto climatizado",
    "Cama",
    "Participação em todas as programações durante a convenção jovem",
    "Momento de lazer",
]

CONTATO_PAGAMENTO = "+55 92 8459-6369"
CONTATO_PAGAMENTO_TEXTO = "Número de contato do pagamento de inscrição"

PIX_KEY_CNPJ = "17.739.576/0001-78"
PIX_QR_STATIC = "img/qr_pix.jpeg"  # dentro do /static

# helpers rápidos


def is_pix(reg: Registration) -> bool:
    return (reg.payment_type or "").lower() == "pix"


def can_admin():
    return getattr(current_user, "can_access_admin", False)


def can_review():
    return getattr(current_user, "can_review_payments", False)


def inscricoes_suspensas() -> bool:
    return os.getenv("INSCRICOES_SUSPENSAS", "0") == "1"


def inscricoes_abertas() -> bool:
    return os.getenv("INSCRICOES_ABERTAS", "0") == "1"


def get_setting(key: str, default: str = "") -> str:
    s = AppSetting.query.filter_by(key=key).first()
    return (s.value if s else default) or default


def inscricoes_status() -> str:
    return get_setting("INSCRICOES_STATUS", "embreve").strip().lower()


# ======================= PIX =======================
PIX_KEY = "17739576000178"


@app.route("/pix/qr/<int:n>")
@login_required
def pix_qr_n(n):
    reg = Registration.query.filter_by(user_id=current_user.id).first_or_404()

    if n not in (1, 2, 3):
        abort(400)

    lot = Decimal(reg.lot_value_cents or 18000) / Decimal(100)

    # parcela inteira (ex: 180/2 = 90) + 0.09
    parcela = (lot / Decimal(n)).quantize(Decimal("0"),
                                          rounding=ROUND_DOWN) + Decimal("0.09")
    parcela = parcela.quantize(Decimal("0.00"))  # garante 2 casas

    txid = f"R{reg.id}N{n}"[:25]  # alfanumérico

    payload = build_pix_payload(
        pix_key=PIX_KEY,
        merchant_name="CONVENCAO AMAZONICA",
        merchant_city="MANAUS",
        amount=str(parcela),   # <-- manda como string "90.09"
        txid=txid
    )

    img = qrcode.make(payload)
    buf = BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return Response(buf.getvalue(), mimetype="image/png")


PIX_PAYLOADS = {
    1: "00020126500014BR.GOV.BCB.PIX0128convencaoamazonica@gmail.com5204000053039865406180.095802BR5925CONVENCAO REGIONAL AMAZON6006MANAUS62250521nNy6i6O8IW9s02P2H3xUy6304452F",
    2: "00020126500014BR.GOV.BCB.PIX0128convencaoamazonica@gmail.com520400005303986540590.095802BR5925CONVENCAO REGIONAL AMAZON6006MANAUS622605222naahyV62ub2ssFa6pm5z763048CD0",
    3: "00020126500014BR.GOV.BCB.PIX0128convencaoamazonica@gmail.com520400005303986540560.095802BR5925CONVENCAO REGIONAL AMAZON6006MANAUS622605223KP4V0Sju0cQ0txYqAWnCu6304E6C0",
}


@app.route("/pix/copia-cola/<int:n>")
@login_required
def pix_copia_cola(n):
    payload = PIX_PAYLOADS.get(n)
    if not payload:
        abort(400)
    return jsonify({"payload": payload})

# ======================= ROUTES =======================


@app.route("/")
def landing():
    v = (request.args.get("v") or "a").lower()

    allowed = {
        "a": "landing_a.html",
        "b": "landing_b.html",
        "c": "landing_c.html",
        "d": "landing_d.html",
        "e": "landing_e.html",
        "f": "landing_f.html",
    }

    tpl = allowed.get(v, "landing_a.html")
    return render_template(
        tpl,
        pix_msg=PIX_PADRAO_MSG,
        criancas_msg=CRIANCAS_MSG,
        inclui_itens=INCLUI_ITENS,
        contato_pagamento=CONTATO_PAGAMENTO,
        contato_pagamento_texto=CONTATO_PAGAMENTO_TEXTO,
        inscricoes_status=inscricoes_status(),
    )


@app.route("/inscricao", methods=["GET", "POST"])
def inscricao():
    st = inscricoes_status()

    if st == "embreve":
        return render_template("inscricoes_em_breve.html",
                               contato_pagamento=CONTATO_PAGAMENTO,
                               contato_pagamento_texto=CONTATO_PAGAMENTO_TEXTO)

    if st == "suspensas":
        return render_template("inscricoes_suspensas.html",
                               contato_pagamento=CONTATO_PAGAMENTO,
                               contato_pagamento_texto=CONTATO_PAGAMENTO_TEXTO)

    # abertas
    form = RegisterAndSignupForm()

    if form.validate_on_submit():
        if form.has_kids_u5.data == "sim" and not (form.kids_u5_names.data or "").strip():
            flash("Informe o nome do(a) filho(a) (5 anos ou menos).", "warning")
            return render_template("inscricao.html", form=form)
        email = form.email.data.strip().lower()

        # evita duplicar conta
        if User.query.filter_by(email=email).first():
            flash(
                "Já existe uma conta com esse e-mail. Faça login para acompanhar.", "warning")
            return redirect(url_for("login"))

        # cria usuário
        user = User(email=email)
        user.set_password(form.password.data)

        # regra de status inicial
        status = "AGUARDANDO_CONFIRMACAO"
        status_msg = "Aguardando confirmação do pagamento."

        reg = Registration(
            user=user,
            full_name=form.full_name.data.strip(),
            cpf=form.cpf.data.strip(),
            phone=form.phone.data.strip(),
            iap_local=form.iap_local.data.strip(),
            transport=form.transport.data,
            payment_type=form.payment_type.data,
            installments=int(form.installments.data),
            lot_name="1_LOTE",
            lot_value_cents=18000,
            status=status,
            status_message=status_msg,

            # ===== NOVOS CAMPOS =====
            age=form.age.data,
            has_kids_u5=(form.has_kids_u5.data == "sim"),
            kids_u5_names=(form.kids_u5_names.data or "").strip() or None,
            is_church_member=(form.is_church_member.data == "sim"),
            agree_no_refund=bool(form.agree_no_refund.data),
        )

        database.session.add(user)
        database.session.add(reg)
        database.session.add(
            AuditLog(action="user_signup_and_register", details=f"email={email}"))
        database.session.commit()

        login_user(user)
        flash("Inscrição criada! Agora você pode acompanhar o status no painel.", "success")
        return redirect(url_for("painel"))

    return render_template("inscricao.html", form=form)


@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("painel"))
    form = LoginForm()
    if form.validate_on_submit():
        email = form.email.data.strip().lower()
        user = User.query.filter_by(email=email).first()

        if not user or not user.check_password(form.password.data):
            flash("E-mail ou senha inválidos.", "danger")
            return render_template("login.html", form=form)

        login_user(user)

        if getattr(user, "must_change_password", False):
            flash("Por segurança, você precisa definir uma nova senha.", "warning")
            return redirect(url_for("change_password"))
        flash("Bem-vindo!", "success")
        return redirect(url_for("painel"))

    return render_template("login.html", form=form)


@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash("Você saiu da sua conta.", "info")
    return redirect(url_for("landing"))


@app.route("/minha-senha", methods=["GET", "POST"])
@login_required
def change_password():
    form = ChangePasswordForm()
    if form.validate_on_submit():
        current_user.set_password(form.new_password.data)
        current_user.must_change_password = False

        database.session.add(AuditLog(
            actor_user_id=current_user.id,
            action="user_change_password",
            details=f"user_id={current_user.id}"
        ))
        database.session.commit()

        flash("Senha atualizada com sucesso!", "success")
        return redirect(url_for("painel"))

    return render_template("change_password.html", form=form)


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    form = LoginForm()

    if form.validate_on_submit():
        email = form.email.data.strip().lower()
        user = User.query.filter_by(email=email).first()

        if not user or not user.check_password(form.password.data):
            flash("E-mail ou senha inválidos.", "danger")
            return render_template("admin/login.html", form=form)

        login_user(user)

        # trava: só entra se tiver acesso ao admin
        if not getattr(user, "can_access_admin", False):
            logout_user()
            flash("Você não tem permissão para acessar o Admin.", "warning")
            return redirect(url_for("login"))

        flash("Acesso administrativo liberado.", "success")
        return redirect(url_for("admin_home"))

    return render_template("admin/login.html", form=form)


@app.route("/painel")
@login_required
def painel():
    reg = current_user.registration  # você já usa 1:1

    total_regs = Registration.query.count()
    lot_info = get_current_lot_info(total_regs)

    # preço "oficial de identificação", sempre terminando em ,09
    price_id = lot_info["price"]  # já vem com .09 se você colocar no env assim
    v1 = price_id
    v2 = split_installments(price_id, 2)[0]  # mesma parcela repetida
    v3 = split_installments(price_id, 3)[0]

    credit_link = "https://link.infinitepay.io/senamarcos/VC1DLTAtUg-7EMh4AJ0qL-180,00"
    return render_template(
        "painel.html",
        reg=reg,
        lot_info=lot_info,
        pix_prices={
            "v1": v1,
            "v2": v2,
            "v3": v3,
        }, credit_link=credit_link,
        pix_payloads=PIX_PAYLOADS,
    )


@app.route("/comprovante")
@login_required
def enviar_comprovante():
    # redireciona pro WhatsApp (se quiser)
    return redirect("https://wa.me/559284596369")

# =======================
# ADMIN
# =======================


@app.route("/admin")
@admin_required
def admin_home():
    # KPIs principais
    total = database.session.query(func.count(Registration.id)).scalar() or 0

    aguardando = (
        database.session.query(func.count(Registration.id))
        .filter(Registration.status == "AGUARDANDO_CONFIRMACAO")
        .scalar()
        or 0
    )

    confirmadas = (
        database.session.query(func.count(Registration.id))
        .filter(Registration.status == "CONFIRMADA")
        .scalar()
        or 0
    )

    negadas = (
        database.session.query(func.count(Registration.id))
        .filter(Registration.status == "NEGADA")
        .scalar()
        or 0
    )

    # Extra: quantas inscrições Pix estão aguardando e já enviaram comprovante
    pendentes_comprovante = (
        database.session.query(func.count(Registration.id))
        .filter(
            Registration.payment_type == "pix",
            Registration.status == "AGUARDANDO_CONFIRMACAO",
            Registration.proof_file_path.isnot(None),
        )
        .scalar()
        or 0
    )

    # Extra: últimas 8 inscrições (para visão rápida)
    ultimas = (
        Registration.query
        .order_by(Registration.created_at.desc())
        .limit(8)
        .all()
    )

    return render_template(
        "admin/home.html",
        kpi_total=total,
        kpi_aguardando=aguardando,
        kpi_confirmadas=confirmadas,
        kpi_negadas=negadas,
        kpi_pendentes_comprovante=pendentes_comprovante,
        ultimas=ultimas,
    )


@app.route("/admin/relatorio-inscritos.xlsx")
@login_required
@admin_required
def admin_relatorio_inscritos_xlsx():
    """
    Relatório completasso em Excel.
    Suporta filtros opcionais via querystring:
      - ?status=CONFIRMADA
      - ?payment_type=pix
      - ?from=2026-02-01&to=2026-02-10 (YYYY-MM-DD)
    """
    q = Registration.query

    status = (request.args.get("status") or "").strip()
    if status:
        q = q.filter(Registration.status == status)

    payment_type = (request.args.get("payment_type") or "").strip()
    if payment_type:
        q = q.filter(Registration.payment_type == payment_type)

    dt_from = (request.args.get("from") or "").strip()
    dt_to = (request.args.get("to") or "").strip()

    def parse_date(s: str):
        try:
            return datetime.strptime(s, "%Y-%m-%d")
        except Exception:
            return None

    d1 = parse_date(dt_from)
    d2 = parse_date(dt_to)
    if d1:
        q = q.filter(Registration.created_at >= d1)
    if d2:
        # inclui o dia final (até 23:59:59)
        q = q.filter(Registration.created_at < datetime(
            d2.year, d2.month, d2.day, 23, 59, 59))

    regs = q.order_by(Registration.created_at.desc()).all()

    # ===== Helpers =====
    def fmt_dt(dt):
        if not dt:
            return ""
        return dt.strftime("%d/%m/%Y %H:%M")

    def yn(b):
        return "SIM" if b else "NÃO"

    def money_from_cents(cents):
        return float((cents or 0) / 100.0)

    def safe_text(v):
        return (v or "").strip() if isinstance(v, str) else (v if v is not None else "")

    def count_by(items, key_fn):
        d = {}
        for it in items:
            k = key_fn(it)
            d[k] = d.get(k, 0) + 1
        return d

    def sort_count_dict(d):
        return sorted(d.items(), key=lambda x: (-x[1], str(x[0])))

    # ===== Workbook styles =====
    wb = Workbook()

    # remove sheet padrão
    wb.remove(wb.active)

    title_font = Font(bold=True, size=14)
    h_font = Font(bold=True, size=11, color="FFFFFF")
    h_fill = PatternFill("solid", fgColor="111827")  # quase preto (dark)
    thin = Side(style="thin", color="2D3748")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header_row(ws, row=1):
        for col in range(1, ws.max_column + 1):
            c = ws.cell(row=row, column=col)
            c.font = h_font
            c.fill = h_fill
            c.alignment = Alignment(vertical="center")
            c.border = border

    def apply_table_style(ws, start_row, start_col, end_row, end_col, freeze_panes="A2", auto_filter=True):
        ws.freeze_panes = freeze_panes
        if auto_filter:
            ws.auto_filter.ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                ws.cell(row=r, column=c).border = border
                ws.cell(row=r, column=c).alignment = Alignment(
                    vertical="top", wrap_text=True)

    def set_widths(ws, widths: dict):
        for col_letter, w in widths.items():
            ws.column_dimensions[col_letter].width = w

    # ===== 1) Resumo =====
    ws_sum = wb.create_sheet("Resumo")
    ws_sum["A1"] = "Relatório de Inscrições — Tempo de Resplandecer"
    ws_sum["A1"].font = title_font

    ws_sum["A3"] = "Gerado em:"
    ws_sum["B3"] = datetime.now().strftime("%d/%m/%Y %H:%M")

    ws_sum["A4"] = "Filtros:"
    filtros = []
    if status:
        filtros.append(f"status={status}")
    if payment_type:
        filtros.append(f"payment_type={payment_type}")
    if dt_from:
        filtros.append(f"from={dt_from}")
    if dt_to:
        filtros.append(f"to={dt_to}")
    ws_sum["B4"] = ", ".join(filtros) if filtros else "—"

    total = len(regs)
    by_status = count_by(regs, lambda r: safe_text(r.status))
    by_pay = count_by(
        regs, lambda r: "Pix" if r.payment_type == "pix" else "Crédito")
    by_inst = count_by(regs, lambda r: f"{int(r.installments or 1)}x")
    by_kids = count_by(regs, lambda r: "SIM" if r.has_kids_u5 else "NÃO")
    by_transport = count_by(
        regs, lambda r: "Ônibus" if r.transport == "onibus" else "Carro")
    by_proof = count_by(
        [r for r in regs if r.payment_type == "pix"],
        lambda r: "SIM" if r.proof_file_path else "NÃO"
    )

    # KPIs em cards simples (linhas)
    ws_sum["A6"] = "KPIs"
    ws_sum["A6"].font = Font(bold=True, size=12)

    kpi_rows = [
        ("Total de inscrições", total),
        ("Confirmadas", by_status.get("CONFIRMADA", 0)),
        ("Aguardando confirmação", by_status.get("AGUARDANDO_CONFIRMACAO", 0)),
        ("Negadas", by_status.get("NEGADA", 0)),
        ("Pagamentos Pix", by_pay.get("Pix", 0)),
        ("Pagamentos Crédito", by_pay.get("Crédito", 0)),
        ("Tem filhos até 5 (SIM)", by_kids.get("SIM", 0)),
        ("Tem filhos até 5 (NÃO)", by_kids.get("NÃO", 0)),
        ("Pix com comprovante (SIM)", by_proof.get("SIM", 0)),
        ("Pix sem comprovante (NÃO)", by_proof.get("NÃO", 0)),
    ]

    start = 8
    ws_sum.append(["Métrica", "Valor"])
    ws_sum["A7"] = "Métrica"
    ws_sum["B7"] = "Valor"
    ws_sum["A7"].font = h_font
    ws_sum["B7"].font = h_font
    ws_sum["A7"].fill = h_fill
    ws_sum["B7"].fill = h_fill
    ws_sum["A7"].border = border
    ws_sum["B7"].border = border

    r0 = 8
    for name, val in kpi_rows:
        ws_sum.cell(row=r0, column=1, value=name)
        ws_sum.cell(row=r0, column=2, value=val)
        ws_sum.cell(row=r0, column=1).border = border
        ws_sum.cell(row=r0, column=2).border = border
        r0 += 1

    set_widths(ws_sum, {"A": 36, "B": 16})

    # ===== 2) Inscritos (tabela completa) =====
    ws = wb.create_sheet("Inscritos")

    headers = [
        "ID",
        "Nome",
        "CPF",
        "Telefone",
        "IAP",
        "Transporte",
        "Lote",
        "Valor Lote (R$)",
        "Tipo Pagamento",
        "Parcelas",
        "Status",
        "Mensagem Status",
        "Tem filhos até 5?",
        "Nomes filhos até 5",
        "Idade",
        "Membro da igreja?",
        "Aceitou sem reembolso?",
        "Comprovante Pix enviado?",
        "Comprovante enviado em",
        "Revisado por (user_id)",
        "Revisado em",
        "Nota revisão",
        "Criado em",
        "Atualizado em",
    ]
    ws.append(headers)
    style_header_row(ws, 1)

    for r in regs:
        ws.append([
            r.id,
            r.full_name,
            r.cpf,
            r.phone,
            r.iap_local,
            "Ônibus" if r.transport == "onibus" else "Carro",
            r.lot_name,
            money_from_cents(r.lot_value_cents),
            "Pix" if r.payment_type == "pix" else "Crédito",
            int(r.installments or 1),
            r.status,
            safe_text(r.status_message),
            yn(r.has_kids_u5),
            safe_text(r.kids_u5_names),
            r.age if r.age is not None else "",
            yn(r.is_church_member),
            yn(r.agree_no_refund),
            ("SIM" if r.proof_file_path else "NÃO") if r.payment_type == "pix" else "N/A",
            fmt_dt(r.proof_uploaded_at) if r.payment_type == "pix" else "",
            r.reviewed_by_user_id or "",
            fmt_dt(r.reviewed_at),
            safe_text(r.review_note),
            fmt_dt(r.created_at),
            fmt_dt(r.updated_at),
        ])

    # formata coluna de dinheiro
    for row in range(2, ws.max_row + 1):
        ws[f"H{row}"].number_format = '0.00'

    # widths
    set_widths(ws, {
        "A": 8,   "B": 30, "C": 16, "D": 16, "E": 24, "F": 12,
        "G": 12,  "H": 14, "I": 14, "J": 10, "K": 22, "L": 30,
        "M": 16,  "N": 32, "O": 8,  "P": 18, "Q": 22, "R": 22,
        "S": 18,  "T": 18, "U": 18, "V": 28, "W": 18, "X": 18,
    })

    apply_table_style(ws, 1, 1, ws.max_row, ws.max_column,
                      freeze_panes="A2", auto_filter=True)

    # ===== Abas de agregação =====
    def make_count_sheet(title, pairs, col1="Categoria", col2="Qtd"):
        s = wb.create_sheet(title)
        s.append([col1, col2])
        style_header_row(s, 1)
        for k, v in pairs:
            s.append([k, v])
        set_widths(s, {"A": 36, "B": 10})
        apply_table_style(s, 1, 1, s.max_row, 2,
                          freeze_panes="A2", auto_filter=True)
        return s

    # 3) Status
    make_count_sheet("Status", sort_count_dict(by_status), "Status", "Qtd")

    # 4) Pagamentos (tipo + parcelas)
    ws_pay = wb.create_sheet("Pagamentos")
    ws_pay.append(["Tipo", "Qtd"])
    ws_pay.append(["Pix", by_pay.get("Pix", 0)])
    ws_pay.append(["Crédito", by_pay.get("Crédito", 0)])
    style_header_row(ws_pay, 1)
    apply_table_style(ws_pay, 1, 1, ws_pay.max_row, 2,
                      freeze_panes="A2", auto_filter=False)
    set_widths(ws_pay, {"A": 20, "B": 10})

    # tabela por parcelas
    ws_pay["D1"] = "Parcelas"
    ws_pay["D1"].font = Font(bold=True, size=12)
    ws_pay.append([])  # só pra manter simples (não afeta)
    ws_pay2_row = 3
    ws_pay.cell(row=2, column=4, value="Parcelas").font = h_font
    ws_pay.cell(row=2, column=5, value="Qtd").font = h_font
    ws_pay.cell(row=2, column=4).fill = h_fill
    ws_pay.cell(row=2, column=5).fill = h_fill
    ws_pay.cell(row=2, column=4).border = border
    ws_pay.cell(row=2, column=5).border = border

    for k, v in sort_count_dict(by_inst):
        ws_pay.cell(row=ws_pay2_row, column=4, value=k).border = border
        ws_pay.cell(row=ws_pay2_row, column=5, value=v).border = border
        ws_pay2_row += 1

    set_widths(ws_pay, {"D": 14, "E": 10})

    # 5) Kids U5
    make_count_sheet("Kids U5", sort_count_dict(
        by_kids), "Tem filhos até 5?", "Qtd")

    # 6) IAP ranking
    by_iap = count_by(regs, lambda r: safe_text(r.iap_local) or "—")
    make_count_sheet("IAP", sort_count_dict(by_iap), "IAP", "Qtd")

    # 7) Lotes
    by_lot = count_by(regs, lambda r: safe_text(r.lot_name) or "—")
    make_count_sheet("Lotes", sort_count_dict(by_lot), "Lote", "Qtd")

    # 8) Transporte
    make_count_sheet("Transporte", sort_count_dict(
        by_transport), "Transporte", "Qtd")

    # ===== Export =====
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="relatorio_inscritos_completo.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/inscricoes")
@payment_reviewer_required
def admin_inscricoes():
    if not can_admin():
        abort(403)

    status = request.args.get("status", "").strip()
    q = request.args.get("q", "").strip()

    # paginação
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 25, type=int)
    per_page = max(10, min(per_page, 100))  # trava entre 10 e 100

    query = Registration.query

    if status:
        query = query.filter(Registration.status == status)

    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            Registration.full_name.ilike(like),
            Registration.cpf.ilike(like),
            Registration.phone.ilike(like),
            Registration.iap_local.ilike(like),
        ))

    query = query.order_by(Registration.created_at.desc())

    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    regs = pagination.items

    return render_template(
        "admin/inscricoes.html",
        regs=regs,
        pagination=pagination,
        status=status,
        q=q,
        per_page=per_page
    )


@app.route("/admin/inscricoes/<int:reg_id>", methods=["GET", "POST"])
@login_required
def admin_inscricao_detalhe(reg_id):
    if not can_review():
        abort(403)

    reg = Registration.query.get_or_404(reg_id)
    form = ReviewRegistrationForm()

    if form.validate_on_submit():
        decision = form.decision.data  # CONFIRMADA | NEGADA
        note = (form.note.data or "").strip()

        reg.status = decision
        if decision == "CONFIRMADA":
            reg.status_message = "Inscrição confirmada. Seja bem-vindo(a)!"
        else:
            reg.status_message = "Inscrição negada. Entre em contato para ajustes."

        reg.review_note = note if note else None
        reg.reviewed_by_user_id = current_user.id
        reg.reviewed_at = datetime.utcnow()

        database.session.add(AuditLog(
            actor_user_id=current_user.id,
            action="review_registration",
            details=f"registration_id={reg.id} decision={decision}"
        ))
        database.session.commit()

        flash("Decisão salva com sucesso!", "success")
        return redirect(url_for("admin_inscricoes", status=request.args.get("status", "")))

    return render_template(
        "admin/inscricao_detalhe.html",
        reg=reg,
        user=reg.user,
        reviewer=reg.reviewer,
        form=form
    )

# =======================
# SUPER USER - PERMISSÕES
# =======================


def gerar_senha_temporaria(tamanho: int = 10) -> str:
    # senha simples e copiável (letras+numeros). Se quiser mais forte, aumento e boto símbolos.
    alfabeto = string.ascii_letters + string.digits
    return "".join(secrets.choice(alfabeto) for _ in range(tamanho))


@app.route("/admin/permissoes", methods=["GET", "POST"])
@super_required
def super_permissoes():
    users = User.query.order_by(User.created_at.desc()).all()
    roles = Role.query.order_by(Role.name.asc()).all()

    if request.method == "POST":
        user_id = int(request.form.get("user_id") or 0)
        # add | remove | reset_password
        action = (request.form.get("action") or "").strip().lower()

        user = User.query.get_or_404(user_id)

        # =========================
        # RESET SENHA
        # =========================
        if action == "reset_password":
            nova_senha = gerar_senha_temporaria(10)
            user.set_password(nova_senha)

            user.must_change_password = True
            user.password_reset_at = datetime.utcnow()

            database.session.add(AuditLog(
                actor_user_id=current_user.id,
                action="super_reset_password",
                details=f"user_id={user_id}"
            ))
            database.session.commit()

            # abre modal no front
            return redirect(url_for("super_permissoes", pw=nova_senha, email=user.email))

        # =========================
        # ADD / REMOVE ROLE
        # =========================
        role_id = int(request.form.get("role_id") or 0)
        role = Role.query.get_or_404(role_id)

        if action == "add" and role not in user.roles:
            user.roles.append(role)
        elif action == "remove" and role in user.roles:
            user.roles.remove(role)

        database.session.add(AuditLog(
            actor_user_id=current_user.id,
            action="super_update_role",
            details=f"user_id={user_id} role_id={role_id} action={action}"
        ))
        database.session.commit()

        flash("Permissões atualizadas.", "success")
        return redirect(url_for("super_permissoes"))

    return render_template("admin/permissoes.html", users=users, roles=roles)


# =======================
# BOOTSTRAP ROLES (rodar 1x)
# =======================
@app.cli.command("seed_roles")
def seed_roles():
    """
    flask seed_roles
    """
    def upsert(name, **kwargs):
        r = Role.query.filter_by(name=name).first()
        if not r:
            r = Role(name=name, **kwargs)
            database.session.add(r)
        else:
            for k, v in kwargs.items():
                setattr(r, k, v)

    upsert("SUPER", is_super=True, can_access_admin=True,
           can_review_payments=True)
    upsert("ADMIN", is_super=False, can_access_admin=True,
           can_review_payments=True)
    upsert("REVISOR_PAGAMENTOS", is_super=False,
           can_access_admin=True, can_review_payments=True)
    database.session.commit()
    print("Roles criadas/atualizadas.")


@app.cli.command("make_super")
def make_super():
    """
    Uso:
      flask make_super admin@teste.com
    """
    import sys

    if len(sys.argv) < 3:
        print("Uso: flask make_super email@dominio.com")
        return

    email = sys.argv[2].strip().lower()
    user = User.query.filter_by(email=email).first()
    if not user:
        print(f"Usuário não encontrado: {email}")
        return

    role_super = Role.query.filter_by(name="SUPER").first()
    if not role_super:
        print("Role SUPER não existe. Rode: flask seed_roles")
        return

    if role_super not in user.roles:
        user.roles.append(role_super)

    database.session.add(AuditLog(
        actor_user_id=None,
        action="cli_make_super",
        details=f"email={email}"
    ))
    database.session.commit()

    print(f"OK! {email} agora é SUPER.")


def generate_temp_password(length=12) -> str:
    alphabet = string.ascii_letters + string.digits
    return "".join(secrets.choice(alphabet) for _ in range(length))


@app.route("/admin/users/<int:user_id>/reset_password", methods=["POST"])
@super_required
def admin_reset_password(user_id):
    user = User.query.get_or_404(user_id)

    temp_pass = generate_temp_password(12)
    user.set_password(temp_pass)
    user.must_change_password = True
    user.password_reset_at = datetime.utcnow()

    database.session.add(AuditLog(
        actor_user_id=current_user.id,
        action="admin_reset_password",
        details=f"user_id={user.id} email={user.email}"
    ))
    database.session.commit()

    # Renderiza uma tela que mostra a senha UMA VEZ
    return render_template("admin/reset_password_result.html", user=user, temp_pass=temp_pass)


@app.route("/admin/config/inscricoes", methods=["GET", "POST"])
@super_required
def admin_config_inscricoes():
    key = "INSCRICOES_STATUS"
    current = get_setting(key, "embreve")

    if request.method == "POST":
        new_status = (request.form.get("status") or "").strip().lower()
        if new_status not in ("abertas", "suspensas", "embreve"):
            flash("Status inválido.", "danger")
            return redirect(url_for("admin_config_inscricoes"))

        s = AppSetting.query.filter_by(key=key).first()
        if not s:
            s = AppSetting(key=key, value=new_status)
            database.session.add(s)
        else:
            s.value = new_status

        database.session.add(AuditLog(
            actor_user_id=current_user.id,
            action="update_inscricoes_status",
            details=f"status={new_status}"
        ))
        database.session.commit()

        flash("Status das inscrições atualizado!", "success")
        return redirect(url_for("admin_config_inscricoes"))

    return render_template("admin/config_inscricoes.html", current=current)
